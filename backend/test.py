from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

d = {'room1': [[{'s1': 1, 's2': 2}, {'s1': 3, 's2': 4}, {'s1': 5, 's2': 6}], 
                [{'s1': 7, 's2': 8}, {'s1': 9, 's2': 10}, {'s1': 11, 's2': 12}], 
                [{'s1': 13, 's2': 14}, {'s1': 15, 's2': 16}, {'s1': 17, 's2': 18}]], 
    'room2': [[{'s1': 1, 's2': 2}, {'s1': 3, 's2': 4}, {'s1': 5, 's2': 6}], 
                [{'s1': 7, 's2': 8}, {'s1': 9, 's2': 10}, {'s1': 11, 's2': 12}], 
                [{'s1': 13, 's2': 14}, {'s1': 15, 's2': 16}, {'s1': 17, 's2': 18}]]
}
# for rooms in d:
    # print('room')
    # print(d[rooms])
    # df = pd.DataFrame(d[rooms])
    # print(df)
    # for rows in d[rooms]:
    #     df = pd.DataFrame(rows)
    #     print(df)
                
def build_room_sheet(ws, room_name, rows):
    max_seats = max(len(row) for row in rows)
    total_columns = max_seats * 2  # s1 and s2 occupy separate columns
    arrow_banner = "^" * (max(5, max_seats * 3))

    def merge_and_set(row, value, font_size=14, bold=True):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
        cell = ws.cell(row=row, column=1, value=value)
        cell.font = Font(size=font_size, bold=bold)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    merge_and_set(1, "Seating Plan", font_size=18)
    merge_and_set(2, "")
    merge_and_set(3, room_name, font_size=16)
    merge_and_set(4, "")
    merge_and_set(5, f"{arrow_banner}  Black Board  {arrow_banner}", font_size=11, bold=False)

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    data_start_row = 7

    for row_offset, row in enumerate(rows):
        excel_row = data_start_row + row_offset
        ws.row_dimensions[excel_row].height = 32
        for seat_idx in range(1, max_seats + 1):
            s1_col = (seat_idx - 1) * 2 + 1
            s2_col = s1_col + 1

            for col in (s1_col, s2_col):
                cell = ws.cell(row=excel_row, column=col)
                ws.column_dimensions[get_column_letter(col)].width = 14
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            if seat_idx <= len(row):
                student = row[seat_idx - 1] or {}
                ws.cell(row=excel_row, column=s1_col).value = student.get("s1", "")
                ws.cell(row=excel_row, column=s2_col).value = student.get("s2", "")


def build_workbook(data: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    first = True

    for room_name, rows in data.items():
        if first:
            ws.title = room_name
            first = False
        else:
            ws = wb.create_sheet(title=room_name)
        build_room_sheet(ws, room_name, rows)

    wb.save(output_path)


build_workbook(d, "output3.xlsx")