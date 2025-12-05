from collections import defaultdict
import math
import os
import re
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter


def _clean_value(value):
    """Convert NaN, None, or pandas NaN to empty string."""
    if value is None:
        return ""
    try:
        # Check for pandas/numpy NaN
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    try:
        # Check for float NaN
        if isinstance(value, float) and math.isnan(value):
            return ""
    except (TypeError, ValueError):
        pass
    # Check for string "nan" or "NaN"
    if isinstance(value, str) and value.lower() in ('nan', 'none', ''):
        return ""
    return value

def _split_roll_and_branch(raw_value: str):
    """Return (roll_no, branch) tuple from a raw string."""
    # Handle NaN, None, or empty values
    if raw_value is None or pd.isna(raw_value):
        return "", ""

    text = str(raw_value).strip()
    if not text or text.lower() == 'nan':
        return "", ""

    if "\n" in text:
        roll, branch = text.split("\n", 1)
    else:
        parts = text.split()
        roll = parts[0]
        branch = " ".join(parts[1:]) if len(parts) > 1 else ""

    return roll.strip(), branch.strip()

def _extract_roll_number(roll_str):
    """Extract numeric part from roll number string.
    Handles formats like '201', '201A', '201-1', etc."""
    if not roll_str:
        return None
    roll_str = str(roll_str).strip()
    # Extract first sequence of digits
    match = re.search(r'\d+', roll_str)
    if match:
        return int(match.group())
    return None

def _find_consecutive_ranges(numbers):
    """Find consecutive number ranges from a list of roll number strings.
    Returns a list of strings like ['201-208', '210-220']"""
    if not numbers:
        return []
    
    # Extract numeric values and filter out None
    numeric_values = []
    for n in numbers:
        num = _extract_roll_number(n)
        if num is not None:
            numeric_values.append(num)
    
    if not numeric_values:
        return []
    
    # Sort and find consecutive ranges
    sorted_nums = sorted(set(numeric_values))  # Remove duplicates and sort
    ranges = []
    start = sorted_nums[0]
    end = sorted_nums[0]
    
    for i in range(1, len(sorted_nums)):
        if sorted_nums[i] == end + 1:
            # Consecutive, extend range
            end = sorted_nums[i]
        else:
            # Gap found, save current range
            if start == end:
                ranges.append(str(start))
            else:
                ranges.append(f"{start}-{end}")
            start = sorted_nums[i]
            end = sorted_nums[i]
    
    # Add the last range
    if start == end:
        ranges.append(str(start))
    else:
        ranges.append(f"{start}-{end}")
    
    return ranges

def upload_students(file):
    df = pd.read_excel(file,
                    sheet_name="main",
                    usecols = ['Roll No. Series-1', 'Roll No. Series-2'],
                    )

    pairs = df.to_dict(orient="records") #[{'Roll No. Series-1': str or nan , 'Roll No. Series-2': '2200970700064\n MBA-II'}]
    return pairs

def upload_rooms(file):
    df = pd.read_excel(file,
                    sheet_name="main",
                    usecols = ['Room No.', 'Row', 'Column'],
                    )

    rooms = df.dropna(how="all").to_dict(orient="records") #[{'Room No.': 'D-104' or nan, 'Row': 8.0 or nan , 'Column': 4.0 or nan}...]
    return rooms

def upload_college_sem(file):
    """Read college name and exam name from Excel file.
    Returns (college_name, exam_name) tuple from the first non-empty record."""
    df = pd.read_excel(file,
                    sheet_name="main",
                    usecols = ['College Name', 'Exam Name'])
    info = df.dropna(how="all").to_dict(orient="records")
    
    if not info:
        return "", ""
    
    first_record = info[0]
    college_name = str(first_record.get("College Name", "")).strip() if first_record.get("College Name") else ""
    exam_name = str(first_record.get("Exam Name", "")).strip() if first_record.get("Exam Name") else ""
    
    return college_name, exam_name

def find_capacity_per_room(rooms: dict) -> dict[Any, Any]:
    room_capacity = {}
    for room in rooms:
        room_no = room['Room No.']
        rows = int(room['Row'] or 0)
        cols = int(room['Column'] or 0)

        room_capacity[room_no] = {
            "rows": rows,
            "cols": cols,
            "capacity": rows * cols
        }
    return room_capacity        #room_capacity = {'D-104': {'rows':8,'cols':4,'capacity':32}...}

def fill_room(pairs: list, room_capacity: dict):
    room_layout = defaultdict(list)  # creates an empty dictionary with values as lists {some_key: []}
    branch_counts_per_room = defaultdict(lambda: defaultdict(int))  # {room_no: {branch: count}}
    pair_idx = 0
    branch_range_per_room = defaultdict(lambda: defaultdict(list))  # {room_no: {branch: [roll_numbers]}}

    for room_no, spec in room_capacity.items():
        rows = int(spec.get("rows", 0) or 0)
        cols = int(spec.get("cols", 0) or 0)

        # build a grid [row][col], but fill column by column so
        # students in the same "current_row" list end up in one column
        grid = [[None for _ in range(cols)] for _ in range(rows)]

        for c in range(cols):
            for r in range(rows):
                if pair_idx >= len(pairs):
                    break
                grid[r][c] = pairs[pair_idx]
                
                # Count branches for this pair
                pair = pairs[pair_idx]
                s1_raw = pair.get("Roll No. Series-1", pair.get("s1", ""))
                s2_raw = pair.get("Roll No. Series-2", pair.get("s2", ""))
                
                roll1, branch1 = _split_roll_and_branch(s1_raw)
                roll2, branch2 = _split_roll_and_branch(s2_raw)
                
                # Track roll numbers for range calculation
                if branch1 and roll1:
                    branch_range_per_room[room_no][branch1].append(roll1)
                    branch_counts_per_room[room_no][branch1] += 1
                if branch2 and roll2:
                    branch_range_per_room[room_no][branch2].append(roll2)
                    branch_counts_per_room[room_no][branch2] += 1
                
                pair_idx += 1
            if pair_idx >= len(pairs):
                break

        # convert grid back to list-of-lists (rows), dropping empty seats
        cleaned_rows = []
        for r in range(rows):
            row_seats = [seat for seat in grid[r] if seat is not None]
            if row_seats:
                cleaned_rows.append(row_seats)

        room_layout[room_no] = cleaned_rows

        if pair_idx >= len(pairs):
            break  # no more students to allocate
    
    # Convert roll number lists to consecutive ranges
    branch_range_per_room_final = {}
    for room_no, branch_rolls in branch_range_per_room.items():
        branch_range_per_room_final[room_no] = {}
        for branch, roll_numbers in branch_rolls.items():
            ranges = _find_consecutive_ranges(roll_numbers)
            if ranges:  # Only add if there are ranges
                branch_range_per_room_final[room_no][branch] = ranges
    
    # Store in the original variable name for access
    branch_range_per_room = branch_range_per_room_final
    
    unallocated = (len(pairs) - pair_idx) * 2
    # Convert defaultdict to regular dict for return
    branch_counts_dict = {room: dict(branches) for room, branches in branch_counts_per_room.items()}
    
    return room_layout,unallocated, branch_counts_dict, branch_range_per_room     # ({'D-104': [[{pair1}, {pair2}, ...], [{pairN}, ...]]}, {'D-104': {'branch1': count, 'branch2': count}})

def build_qpd_sheet(ws, branch_counts_per_room: dict, college_name: str = "", exam_name: str = "", 
                    date: str = "", shift_time: str = "", unallocated: int = 0):
    """
    Build a formatted QPD (Quarterly Progress Distribution) sheet on the given worksheet.
    
    Args:
        ws: openpyxl worksheet object to build the QPD sheet on
        branch_counts_per_room: Dict like {'D-104': {'IT-II': 32, 'MBA-IV': 32}, ...}
        college_name: Name of the college
        exam_name: Name of the exam
        date: Date string (e.g., "04-07-2023")
        shift_time: Shift time (e.g., "10:00-12:00")
        unallocated: Number of unallocated students
    """
    # Build the nested structure: {semester: {branch: {room: count}}}
    qpd = defaultdict(lambda: defaultdict(dict))
    all_rooms = []  # Preserve order from input
    
    for room, branch_counts in branch_counts_per_room.items():
        if room not in all_rooms:
            all_rooms.append(room)
        for branch_sem, count in branch_counts.items():
            # Split branch-semester like 'IT-II' into branch='IT' and semester='II'
            if '-' in branch_sem:
                branch_name, semester = branch_sem.rsplit('-', 1)
                qpd[semester][branch_name][room] = count
    
    # Get all semesters and branches, sorted
    semesters = sorted(qpd.keys())
    all_branches_by_sem = {}
    for sem in semesters:
        all_branches_by_sem[sem] = sorted(qpd[sem].keys())
    
    # Styling - use thick borders for entire table
    thick = Side(border_style="thick", color="000000")
    border = Border(top=thick, bottom=thick, left=thick, right=thick)
    header_border = Border(top=thick, bottom=thick, left=thick, right=thick)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    current_row = 1
    
    # Determine column structure first (needed for header width calculation)
    col = 1
    room_col_left = col
    col += 1
    
    semester_start_cols = {}
    semester_col_ranges = {}
    
    for sem in semesters:
        semester_start_cols[sem] = col
        num_branches = len(all_branches_by_sem[sem])
        semester_col_ranges[sem] = (col, col + num_branches - 1)
        col += num_branches
    
    total_col = col
    col += 1
    room_col_right = col
    table_width = room_col_right
    
    # Header section
    if exam_name:
        title = f"QPD - {exam_name}"
        ws.merge_cells(start_row=current_row, start_column=1, 
                      end_row=current_row, end_column=table_width)
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
    
    if date or shift_time:
        info_text = ""
        if date:
            info_text = f"Date: {date}"
        if shift_time:
            if info_text:
                info_text += f" | Shift: {shift_time}"
            else:
                info_text = f"Shift: {shift_time}"
        
        ws.merge_cells(start_row=current_row, start_column=1, 
                      end_row=current_row, end_column=table_width)
        cell = ws.cell(row=current_row, column=1, value=info_text)
        cell.font = Font(size=12, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 2  # Extra space after header
    
    # Build header rows
    header_row = current_row
    
    # First header row: Semester labels
    cell = ws.cell(row=header_row, column=room_col_left, value="ROOM NO.")
    cell.font = Font(size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = header_border
    
    for sem in semesters:
        start_col, end_col = semester_col_ranges[sem]
        ws.merge_cells(start_row=header_row, start_column=start_col,
                      end_row=header_row, end_column=end_col)
        cell = ws.cell(row=header_row, column=start_col, value=f"{sem} SEM")
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
    
    cell = ws.cell(row=header_row, column=total_col, value="Total")
    cell.font = Font(size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = header_border
    
    cell = ws.cell(row=header_row, column=room_col_right, value="ROOM NO.")
    cell.font = Font(size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = header_border
    
    current_row += 1
    
    # Second header row: Branch names
    for sem in semesters:
        start_col = semester_start_cols[sem]
        for idx, branch in enumerate(all_branches_by_sem[sem]):
            col_idx = start_col + idx
            cell = ws.cell(row=current_row, column=col_idx, value=branch)
            cell.font = Font(size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = header_border
    
    # Empty cells for room columns and total in second header row
    for col_idx in [room_col_left, total_col, room_col_right]:
        cell = ws.cell(row=current_row, column=col_idx, value="")
        cell.border = header_border
    
    current_row += 1
    
    # Data rows - preserve original order from input
    for room in all_rooms:
        row_total = 0
        
        # Left ROOM NO.
        cell = ws.cell(row=current_row, column=room_col_left, value=room)
        cell.font = Font(size=10, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        
        # Semester columns
        for sem in semesters:
            start_col = semester_start_cols[sem]
            for idx, branch in enumerate(all_branches_by_sem[sem]):
                count = qpd[sem].get(branch, {}).get(room, 0)
                if count > 0:
                    row_total += count
                
                col_idx = start_col + idx
                cell = ws.cell(row=current_row, column=col_idx, value=count if count > 0 else "")
                cell.font = Font(size=10, bold=False)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                if count > 0:
                    cell.fill = yellow_fill
        
        # Total
        cell = ws.cell(row=current_row, column=total_col, value=row_total if row_total > 0 else "")
        cell.font = Font(size=10, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        if row_total > 0:
            cell.fill = yellow_fill
        
        # Right ROOM NO.
        cell = ws.cell(row=current_row, column=room_col_right, value=room)
        cell.font = Font(size=10, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        
        current_row += 1
    
    # Summary row
    summary_row = current_row
    current_row += 1
    
    # Summary label
    cell = ws.cell(row=summary_row, column=room_col_left, value="Total")
    cell.font = Font(size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = header_border
    
    grand_total = 0
    
    # Semester totals
    for sem in semesters:
        start_col = semester_start_cols[sem]
        for idx, branch in enumerate(all_branches_by_sem[sem]):
            branch_total = sum(qpd[sem].get(branch, {}).values())
            grand_total += branch_total
            
            col_idx = start_col + idx
            cell = ws.cell(row=summary_row, column=col_idx, value=branch_total if branch_total > 0 else "")
            cell.font = Font(size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = header_border
            if branch_total > 0:
                cell.fill = yellow_fill
    
    # Grand total
    cell = ws.cell(row=summary_row, column=total_col, value=grand_total)
    cell.font = Font(size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = header_border
    if grand_total > 0:
        cell.fill = yellow_fill
    
    cell = ws.cell(row=summary_row, column=room_col_right, value="")
    cell.border = header_border
    
    current_row += 1
    
    # Unallocated row
    if unallocated > 0:
        unallocated_row = current_row
        current_row += 1
        
        # Left ROOM NO. - show "Unallocated"
        cell = ws.cell(row=unallocated_row, column=room_col_left, value="Unallocated")
        cell.font = Font(size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
        
        # Empty cells for all semester/branch columns
        for sem in semesters:
            start_col = semester_start_cols[sem]
            for idx, branch in enumerate(all_branches_by_sem[sem]):
                col_idx = start_col + idx
                cell = ws.cell(row=unallocated_row, column=col_idx, value="")
                cell.border = header_border
        
        # Total column - show unallocated count
        cell = ws.cell(row=unallocated_row, column=total_col, value=unallocated)
        cell.font = Font(size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
        cell.fill = yellow_fill
        
        # Right ROOM NO. - empty
        cell = ws.cell(row=unallocated_row, column=room_col_right, value="")
        cell.border = header_border
    
    # Add header rows at the bottom (branches first, then semester labels)
    bottom_header_row1 = current_row  # Branch names row (first at bottom)
    current_row += 1
    
    # First header row at bottom: Branch names
    for sem in semesters:
        start_col = semester_start_cols[sem]
        for idx, branch in enumerate(all_branches_by_sem[sem]):
            col_idx = start_col + idx
            cell = ws.cell(row=bottom_header_row1, column=col_idx, value=branch)
            cell.font = Font(size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = header_border
    
    # Empty cells for room columns and total in first header row at bottom
    for col_idx in [room_col_left, total_col, room_col_right]:
        cell = ws.cell(row=bottom_header_row1, column=col_idx, value="")
        cell.border = header_border
    
    bottom_header_row2 = current_row  # Semester labels row (second at bottom)
    
    # Second header row at bottom: Semester labels and ROOM NO.
    cell = ws.cell(row=bottom_header_row2, column=room_col_left, value="ROOM NO.")
    cell.font = Font(size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = header_border
    
    for sem in semesters:
        start_col, end_col = semester_col_ranges[sem]
        ws.merge_cells(start_row=bottom_header_row2, start_column=start_col,
                      end_row=bottom_header_row2, end_column=end_col)
        cell = ws.cell(row=bottom_header_row2, column=start_col, value=f"{sem} SEM")
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
    
    cell = ws.cell(row=bottom_header_row2, column=total_col, value="Total")
    cell.font = Font(size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = header_border
    
    cell = ws.cell(row=bottom_header_row2, column=room_col_right, value="ROOM NO.")
    cell.font = Font(size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = header_border
    
    # Set column widths
    ws.column_dimensions[get_column_letter(room_col_left)].width = 12
    ws.column_dimensions[get_column_letter(room_col_right)].width = 12
    ws.column_dimensions[get_column_letter(total_col)].width = 10
    
    for sem in semesters:
        start_col = semester_start_cols[sem]
        for idx, branch in enumerate(all_branches_by_sem[sem]):
            col_idx = start_col + idx
            ws.column_dimensions[get_column_letter(col_idx)].width = 10
    
    # Set row heights
    for row in range(header_row, current_row + 1):
        ws.row_dimensions[row].height = 20


def build_msp_base_sheet(ws, branch_range_per_room: dict):
    """
    Build MSP_BASE (Master Student Plan Base) sheet showing roll number ranges per room and branch.
    
    Args:
        ws: openpyxl worksheet object to build the MSP_BASE sheet on
        branch_range_per_room: Dict like {'D-104': {'IT-II': ['201-208', '210-220'], 'EE-IV': ['401-410']}, ...}
    """
    # Styling - use thick borders throughout
    thick = Side(border_style="thick", color="000000")
    border = Border(top=thick, bottom=thick, left=thick, right=thick)
    header_border = Border(top=thick, bottom=thick, left=thick, right=thick)
    
    current_row = 1
    
    # Headers
    headers = ["Room No.", "Branch", "Student Roll Nos."]
    header_row = current_row
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
    
    current_row += 1
    
    # Data rows - preserve order of rooms (not sorted)
    for room_no in branch_range_per_room.keys():
        branches = branch_range_per_room[room_no]
        if not branches:
            continue
        
        # Get all branches for this room (sorted for consistent display)
        branch_list = sorted(branches.keys())
        num_branches = len(branch_list)
        
        # First branch row
        first_branch_row = current_row
        
        # Room No. cell (will be merged if multiple branches)
        room_cell = ws.cell(row=first_branch_row, column=1, value=room_no)
        room_cell.font = Font(size=10, bold=False)
        room_cell.alignment = Alignment(horizontal="left", vertical="center")
        room_cell.border = border
        
        # Process each branch in this room
        for branch_idx, branch in enumerate(branch_list):
            row_num = first_branch_row + branch_idx
            
            # Branch
            branch_cell = ws.cell(row=row_num, column=2, value=branch)
            branch_cell.font = Font(size=10, bold=False)
            branch_cell.alignment = Alignment(horizontal="left", vertical="center")
            branch_cell.border = border
            
            # Format roll numbers
            ranges = branches[branch]
            formatted_ranges = []
            for r in ranges:
                if '-' in r:
                    # Range format: "201-208" -> "(201 to 208)"
                    start, end = r.split('-', 1)
                    formatted_ranges.append(f"({start} to {end})")
                else:
                    # Single number
                    formatted_ranges.append(r)
            
            roll_nos_text = ", ".join(formatted_ranges)
            roll_cell = ws.cell(row=row_num, column=3, value=roll_nos_text)
            roll_cell.font = Font(size=10, bold=False)
            roll_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            roll_cell.border = border
            
            # Calculate row height based on content length (reduced from before)
            # Excel column width of 80 ≈ 80 characters (varies by font)
            # Estimate wrapped lines: account for text length and wrapping
            text_length = len(roll_nos_text)
            column_width_chars = 80  # Column width in characters
            # Account for comma+space separators (avg 2 chars per range/number)
            # More conservative estimate to ensure all text fits
            chars_per_line = max(60, column_width_chars - 20)  # Account for padding
            estimated_lines = max(1, (text_length + chars_per_line - 1) // chars_per_line)
            # Excel row height: reduced base height per line
            base_height_per_line = 13  # Reduced further
            calculated_height = max(18, estimated_lines * base_height_per_line + 3)  # Reduced padding
            # Cap at reasonable maximum but allow for very long lists
            row_height = min(calculated_height, 250)  # Reduced max further
            ws.row_dimensions[row_num].height = row_height
        
        # Update current_row after processing all branches
        current_row = first_branch_row + num_branches
        
        # Merge Room No. cell if multiple branches
        if num_branches > 1:
            ws.merge_cells(start_row=first_branch_row, start_column=1,
                          end_row=current_row - 1, end_column=1)
            # Reapply border to all cells in merged area
            for r in range(first_branch_row, current_row):
                cell = ws.cell(row=r, column=1)
                cell.border = border
    
    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 15  # Room No.
    ws.column_dimensions[get_column_letter(2)].width = 20   # Branch
    ws.column_dimensions[get_column_letter(3)].width = 80  # Student Roll Nos.
    
    # Set header row height
    ws.row_dimensions[header_row].height = 20


def build_msp_sheet(ws, branch_range_per_room: dict):
    """
    Build MSP (Master Student Plan) sheet showing roll number ranges grouped by branch.
    Structure: Branch -> Student Roll Nos. -> Room No.
    
    Args:
        ws: openpyxl worksheet object to build the MSP sheet on
        branch_range_per_room: Dict like {'D-104': {'IT-II': ['201-208', '210-220'], 'EE-IV': ['401-410']}, ...}
    """
    # Transform data structure: from {room: {branch: ranges}} to {branch: {room: ranges}}
    # Preserve branch order and room order as they appear in the original data
    branch_to_room_ranges = defaultdict(lambda: defaultdict(list))
    branch_order = []
    branch_room_order = defaultdict(list)  # Track room order for each branch
    
    for room_no, branches in branch_range_per_room.items():
        for branch, ranges in branches.items():
            if branch not in branch_order:
                branch_order.append(branch)
            if room_no not in branch_room_order[branch]:
                branch_room_order[branch].append(room_no)
            branch_to_room_ranges[branch][room_no].extend(ranges)
    
    # Styling - use thick borders throughout
    thick = Side(border_style="thick", color="000000")
    border = Border(top=thick, bottom=thick, left=thick, right=thick)
    header_border = Border(top=thick, bottom=thick, left=thick, right=thick)
    
    current_row = 1
    
    # Headers
    headers = ["Branch", "Student Roll Nos.", "Room No."]
    header_row = current_row
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
    
    current_row += 1
    
    # Data rows - grouped by branch (preserve order)
    for branch in branch_order:
        rooms = branch_to_room_ranges[branch]
        room_list = branch_room_order[branch]  # Preserve room order as they first appeared
        num_rooms = len(room_list)
        
        # First room row for this branch
        first_branch_row = current_row
        
        # Branch cell (will be merged if multiple rooms)
        branch_cell = ws.cell(row=first_branch_row, column=1, value=branch)
        branch_cell.font = Font(size=10, bold=False)
        branch_cell.alignment = Alignment(horizontal="center", vertical="center")
        branch_cell.border = border
        
        # Process each room for this branch
        for room_idx, room_no in enumerate(room_list):
            row_num = first_branch_row + room_idx
            
            # Format roll numbers for this branch-room combination
            ranges = rooms[room_no]
            formatted_ranges = []
            for r in ranges:
                if '-' in r:
                    # Range format: "201-208" -> "(201 to 208)"
                    start, end = r.split('-', 1)
                    formatted_ranges.append(f"({start} to {end})")
                else:
                    # Single number
                    formatted_ranges.append(r)
            
            roll_nos_text = ", ".join(formatted_ranges)
            roll_cell = ws.cell(row=row_num, column=2, value=roll_nos_text)
            roll_cell.font = Font(size=10, bold=False)
            roll_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            roll_cell.border = border
            
            # Room No.
            room_cell = ws.cell(row=row_num, column=3, value=room_no)
            room_cell.font = Font(size=10, bold=False)
            room_cell.alignment = Alignment(horizontal="center", vertical="center")
            room_cell.border = border
            
            # Calculate row height based on content length
            text_length = len(roll_nos_text)
            column_width_chars = 100  # Column width in characters for roll nos
            chars_per_line = max(70, column_width_chars - 30)  # Account for padding
            estimated_lines = max(1, (text_length + chars_per_line - 1) // chars_per_line)
            # Excel row height: reduced base height per line
            base_height_per_line = 13
            calculated_height = max(18, estimated_lines * base_height_per_line + 3)
            # Cap at reasonable maximum
            row_height = min(calculated_height, 250)
            ws.row_dimensions[row_num].height = row_height
        
        # Update current_row after processing all rooms for this branch
        current_row = first_branch_row + num_rooms
        
        # Merge Branch cell if multiple rooms
        if num_rooms > 1:
            ws.merge_cells(start_row=first_branch_row, start_column=1,
                          end_row=current_row - 1, end_column=1)
            # Reapply border to all cells in merged area
            for r in range(first_branch_row, current_row):
                cell = ws.cell(row=r, column=1)
                cell.border = border
    
    # Set column widths - wider for better content visibility
    ws.column_dimensions[get_column_letter(1)].width = 30  # Branch
    ws.column_dimensions[get_column_letter(2)].width = 120  # Student Roll Nos. (wider for long lists)
    ws.column_dimensions[get_column_letter(3)].width = 18  # Room No.
    
    # Set header row height
    ws.row_dimensions[header_row].height = 25


def generate_qpd(branch_counts_per_room: dict, college_name: str = "", exam_name: str = "", 
                 date: str = "", shift_time: str = "", output_path: str = "qpd.xlsx", unallocated: int = 0):
    """
    Generate a formatted QPD (Question Paper Distribution) Excel file.
    
    Args:
        branch_counts_per_room: Dict like {'D-104': {'IT-II': 32, 'MBA-IV': 32}, ...}
        college_name: Name of the college
        exam_name: Name of the exam
        date: Date string (e.g., "04-07-2023")
        shift_time: Shift time (e.g., "10:00-12:00")
        output_path: Path to save the Excel file
        unallocated: Number of unallocated students
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "QPD"
    
    build_qpd_sheet(ws, branch_counts_per_room, college_name, exam_name, date, shift_time, unallocated)
    
    wb.save(output_path)


def fill_room_row_gap(pairs: list, room_capacity: dict):
    room_layout = defaultdict(list)
    branch_counts_per_room = defaultdict(lambda: defaultdict(int))  # {room_no: {branch: count}}
    pair_idx = 0
    branch_range_per_room = defaultdict(lambda: defaultdict(list))  # {room_no: {branch: [roll_numbers]}}

    for room_no, spec in room_capacity.items():
        rows = int(spec.get("rows", 0) or 0)
        cols = int(spec.get("cols", 0) or 0)
        grid = [[None for _ in range(cols)] for _ in range(rows)]

        for c in range(cols):
            for r in range(rows):
                if r % 2 != 0:  # skip odd rows to keep alternate rows empty
                    continue
                if pair_idx >= len(pairs):
                    break
                grid[r][c] = pairs[pair_idx]
                
                # Count branches for this pair
                pair = pairs[pair_idx]
                s1_raw = pair.get("Roll No. Series-1", pair.get("s1", ""))
                s2_raw = pair.get("Roll No. Series-2", pair.get("s2", ""))
                
                roll1, branch1 = _split_roll_and_branch(s1_raw)
                roll2, branch2 = _split_roll_and_branch(s2_raw)
                
                # Track roll numbers for range calculation
                if branch1 and roll1:
                    branch_range_per_room[room_no][branch1].append(roll1)
                    branch_counts_per_room[room_no][branch1] += 1
                if branch2 and roll2:
                    branch_range_per_room[room_no][branch2].append(roll2)
                    branch_counts_per_room[room_no][branch2] += 1
                
                pair_idx += 1
            if pair_idx >= len(pairs):
                break

        cleaned_rows = []
        for r in range(rows):
            row_seats = [seat for seat in grid[r] if seat is not None]
            # Include all rows (filled or empty) to show skipped alternate rows
            if row_seats:
                cleaned_rows.append(row_seats)
            else:
                # Insert empty row to indicate this row was skipped
                cleaned_rows.append([])

        room_layout[room_no] = cleaned_rows

        if pair_idx >= len(pairs):
            break
    
    # Convert roll number lists to consecutive ranges
    branch_range_per_room_final = {}
    for room_no, branch_rolls in branch_range_per_room.items():
        branch_range_per_room_final[room_no] = {}
        for branch, roll_numbers in branch_rolls.items():
            ranges = _find_consecutive_ranges(roll_numbers)
            if ranges:  # Only add if there are ranges
                branch_range_per_room_final[room_no][branch] = ranges
    
    # Store in the original variable name for access
    branch_range_per_room = branch_range_per_room_final

    unallocated = (len(pairs) - pair_idx) * 2
    # Convert defaultdict to regular dict for return
    branch_counts_dict = {room: dict(branches) for room, branches in branch_counts_per_room.items()}
    
    return room_layout, unallocated, branch_counts_dict, branch_range_per_room

def fill_room_col_gap(pairs: list, room_capacity: dict):
    room_layout = defaultdict(list)
    branch_counts_per_room = defaultdict(lambda: defaultdict(int))  # {room_no: {branch: count}}
    pair_idx = 0
    branch_range_per_room = defaultdict(lambda: defaultdict(list))  # {room_no: {branch: [roll_numbers]}}

    for room_no, spec in room_capacity.items():
        rows = int(spec.get("rows", 0) or 0)
        cols = int(spec.get("cols", 0) or 0)
        grid = [[None for _ in range(cols)] for _ in range(rows)]

        for c in range(cols):
            if c % 2 != 0:  # skip odd columns to leave a gap between columns
                continue
            for r in range(rows):
                if pair_idx >= len(pairs):
                    break
                grid[r][c] = pairs[pair_idx]
                
                # Count branches for this pair
                pair = pairs[pair_idx]
                s1_raw = pair.get("Roll No. Series-1", pair.get("s1", ""))
                s2_raw = pair.get("Roll No. Series-2", pair.get("s2", ""))
                
                roll1, branch1 = _split_roll_and_branch(s1_raw)
                roll2, branch2 = _split_roll_and_branch(s2_raw)
                
                # Track roll numbers for range calculation
                if branch1 and roll1:
                    branch_range_per_room[room_no][branch1].append(roll1)
                    branch_counts_per_room[room_no][branch1] += 1
                if branch2 and roll2:
                    branch_range_per_room[room_no][branch2].append(roll2)
                    branch_counts_per_room[room_no][branch2] += 1
                
                pair_idx += 1
            if pair_idx >= len(pairs):
                break

        cleaned_rows = []
        for r in range(rows):
            # Include all columns (filled or empty) to show skipped alternate columns
            row_seats = grid[r]  # Keep all columns, including None for skipped columns
            # Only include rows that have at least one filled seat
            if any(seat is not None for seat in row_seats):
                cleaned_rows.append(row_seats)

        room_layout[room_no] = cleaned_rows

        if pair_idx >= len(pairs):
            break
    
    # Convert roll number lists to consecutive ranges
    branch_range_per_room_final = {}
    for room_no, branch_rolls in branch_range_per_room.items():
        branch_range_per_room_final[room_no] = {}
        for branch, roll_numbers in branch_rolls.items():
            ranges = _find_consecutive_ranges(roll_numbers)
            if ranges:  # Only add if there are ranges
                branch_range_per_room_final[room_no][branch] = ranges
    
    # Store in the original variable name for access
    branch_range_per_room = branch_range_per_room_final

    unallocated = (len(pairs) - pair_idx) * 2
    # Convert defaultdict to regular dict for return
    branch_counts_dict = {room: dict(branches) for room, branches in branch_counts_per_room.items()}
    
    return room_layout, unallocated, branch_counts_dict, branch_range_per_room


def build_room_sheet(ws, room_name: str, rows: list, college_name: str = "", exam_name: str = "", branch_counts: dict = None):
    if not rows:
        return

    max_seats = max(len(row) for row in rows)
    total_columns = max(1, max_seats * 2)  # s1 and s2 occupy separate columns
    arrow_banner = "^" * (max(5, total_columns * 2))
    
    # Use provided branch_counts or empty dict if not provided
    if branch_counts is None:
        branch_counts = {}

    # Calculate required width for college name (font size 20, bold)
    # Excel column width: 1 unit ≈ 1 character at default font size
    # For font size 20, we need approximately: len(text) * (20/11) * 1.2 (for bold)
    base_column_width = 18  # Default width for seating columns
    column_width = base_column_width  # Will be adjusted if needed
    
    if college_name:
        college_name_clean = _clean_value(college_name) or ""
        if college_name_clean:
            # Estimate width needed: account for larger font (20pt vs 11pt default) and bold
            # Font size 20 is ~1.8x larger than default 11pt, plus 1.3 factor for bold and spacing
            required_width = len(college_name_clean) * (20 / 11) * 1.3
            current_total_width = total_columns * base_column_width
            
            # If college name needs more width, adjust column widths
            if required_width > current_total_width:
                # Calculate new column width to accommodate college name
                column_width = max(base_column_width, required_width / total_columns)
    
    # Set all columns to the calculated width (before displaying college name)
    for col in range(1, total_columns + 1):
        ws.column_dimensions[get_column_letter(col)].width = column_width

    def merge_and_set(row_idx, value, font_size=14, bold=True):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_columns)
        cell = ws.cell(row=row_idx, column=1, value=_clean_value(value))
        cell.font = Font(size=font_size, bold=bold)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 1
    
    # Display college name in big font (no blank line after)
    if college_name:
        merge_and_set(current_row, college_name, font_size=20, bold=True)
        current_row += 1
    
    # Display exam name in big font (no blank line after)
    if exam_name:
        merge_and_set(current_row, exam_name, font_size=20, bold=True)
        current_row += 1
    
    # Display 'Seating Plan' heading (no blank line after)
    merge_and_set(current_row, "Seating Plan", font_size=18)
    current_row += 1
    
    # Display room name (no blank line after)
    merge_and_set(current_row, room_name, font_size=16)
    current_row += 1

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    # data_start_row starts right after the room name (blackboard will be first row of table)
    data_start_row = current_row + 1
    
    # Add blackboard heading as first row of the table
    blackboard_row = data_start_row
    ws.row_dimensions[blackboard_row].height = 36
    ws.merge_cells(start_row=blackboard_row, start_column=1, end_row=blackboard_row, end_column=total_columns)
    blackboard_cell = ws.cell(row=blackboard_row, column=1, value=f"{arrow_banner}  Black Board  {arrow_banner}")
    blackboard_cell.font = Font(size=11, bold=False)
    blackboard_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Apply border to the merged cell (apply to all cells in merged range for proper display)
    for col in range(1, total_columns + 1):
        ws.cell(row=blackboard_row, column=col).border = border
    
    # Adjust data_start_row to start after blackboard row
    data_start_row = blackboard_row + 1

    for row_offset, row in enumerate(rows):
        excel_row = data_start_row + row_offset
        ws.row_dimensions[excel_row].height = 36

        for seat_idx in range(1, max_seats + 1):
            s1_col = (seat_idx - 1) * 2 + 1
            s2_col = s1_col + 1

            for col in (s1_col, s2_col):
                cell = ws.cell(row=excel_row, column=col)
                # Don't overwrite column width - use the width already set for college name
                # ws.column_dimensions[get_column_letter(col)].width is already set above
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            if seat_idx <= len(row):
                student = row[seat_idx - 1] or {}
                s1_raw = student.get("Roll No. Series-1", student.get("s1", ""))
                s2_raw = student.get("Roll No. Series-2", student.get("s2", ""))
                
                # Clean NaN values before processing
                s1_raw = _clean_value(s1_raw) if s1_raw else ""
                s2_raw = _clean_value(s2_raw) if s2_raw else ""

                roll1, branch1 = _split_roll_and_branch(s1_raw)
                roll2, branch2 = _split_roll_and_branch(s2_raw)

                s1_value = "\n".join(filter(None, [roll1, branch1]))
                s2_value = "\n".join(filter(None, [roll2, branch2]))
                ws.cell(row=excel_row, column=s1_col).value = s1_value if s1_value else ""
                ws.cell(row=excel_row, column=s2_col).value = s2_value if s2_value else ""

    if branch_counts:
        summary_start = data_start_row + len(rows) + 2
        # Use columns after the seating table to avoid conflicts
        # Place summary in a dedicated area (columns 1-2, but ensure proper width)
        summary_col1 = 1
        summary_col2 = 2
        
        # Header row
        header_row = summary_start
        name_header = ws.cell(header_row, summary_col1, "Branch Name")
        name_header.font = Font(bold=True, size=12)
        name_header.alignment = Alignment(horizontal="left", vertical="center")
        
        count_header = ws.cell(header_row, summary_col2, "No. of Students")
        count_header.font = Font(bold=True, size=12)
        count_header.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ensure column widths are adequate for summary (only if not already set wider)
        if summary_col1 <= total_columns:
            current_width = ws.column_dimensions[get_column_letter(summary_col1)].width or 18
            ws.column_dimensions[get_column_letter(summary_col1)].width = max(current_width, 20)
        else:
            ws.column_dimensions[get_column_letter(summary_col1)].width = 20
            
        if summary_col2 <= total_columns:
            current_width = ws.column_dimensions[get_column_letter(summary_col2)].width or 18
            ws.column_dimensions[get_column_letter(summary_col2)].width = max(current_width, 18)
        else:
            ws.column_dimensions[get_column_letter(summary_col2)].width = 18
        
        # Data rows - each branch gets its own row
        for idx, (branch, count) in enumerate(branch_counts.items(), start=1):
            data_row = summary_start + idx
            name_cell = ws.cell(data_row, summary_col1, _clean_value(branch))
            name_cell.alignment = Alignment(horizontal="left", vertical="center")
            name_cell.font = Font(size=11)
            
            count_cell = ws.cell(data_row, summary_col2, _clean_value(count))
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            count_cell.font = Font(size=11)


def build_workbook(room_layout: dict, output_path: str = "C:/Users/Ankita/OneDrive/Desktop/sample.xlsx", college_name: str = "", exam_name: str = "", 
                  branch_counts_per_room: dict = None, unallocated: int = 0, date: str = "", shift_time: str = "",
                  branch_range_per_room: dict = None):
    """
    Build or update an Excel workbook with QPD/MSP and room-wise layouts.

    - If `output_path` already exists, it is loaded and **kept intact**.
      New sheets are inserted **after the 'main' sheet** in that workbook.
    - If `output_path` does not exist, a new workbook is created.
    """
    # Load existing workbook if it exists, otherwise create a new one
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        # Remove default sheet in a brand-new workbook
        if wb.worksheets:
            wb.remove(wb.active)

    # Determine insertion index: right after 'main' if present, else at the end
    sheet_names = wb.sheetnames
    if "main" in sheet_names:
        insert_index = sheet_names.index("main") + 1
    else:
        insert_index = len(wb.worksheets)

    # Helper to create/replace a sheet at the current insert_index
    def create_or_replace_sheet(title: str):
        nonlocal insert_index
        if title in wb.sheetnames:
            wb.remove(wb[title])
        ws_local = wb.create_sheet(title=title, index=insert_index)
        insert_index += 1
        return ws_local

    # Create QPD sheet first if branch_counts_per_room is provided
    if branch_counts_per_room:
        qpd_ws = create_or_replace_sheet("QPD")
        build_qpd_sheet(qpd_ws, branch_counts_per_room, college_name, exam_name, date, shift_time, unallocated)
    
    # Create MSP_BASE and MSP sheets if branch_range_per_room is provided
    if branch_range_per_room:
        msp_base_ws = create_or_replace_sheet("MSP_BASE")
        build_msp_base_sheet(msp_base_ws, branch_range_per_room)

        msp_ws = create_or_replace_sheet("MSP")
        build_msp_sheet(msp_ws, branch_range_per_room)
    
    # Create room layout sheets (one per room), after the analytic sheets
    for room_name, rows in room_layout.items():
        # Replace existing sheet with same room name, if any
        ws = create_or_replace_sheet(room_name)
        # Get branch counts for this room if provided
        branch_counts = branch_counts_per_room.get(room_name, {}) if branch_counts_per_room else {}
        build_room_sheet(ws, room_name, rows, college_name, exam_name, branch_counts)

    wb.save(output_path)

    print(f"Workbook created: {output_path}")
    print(f"Workbook created with {len(wb.sheetnames)} sheets")
    print(f"Sheet names: {wb.sheetnames}")


if __name__ == "__main__":
    ### CHANGE PATH
    with open("C:/Users/Ankita/OneDrive/Desktop/sample.xlsx", "rb") as f:
        pairs = upload_students(f)
        rooms = upload_rooms(f)
        college_name, exam_name = upload_college_sem(f)
        room_capacity = find_capacity_per_room(rooms)

        ### UNCOMMENT ANY ONE FUNCTION TO GENERATE DIFFERENT FORMATS
        
        room_layout, unallocated, branch_counts_per_room, branch_range_per_room = fill_room(pairs, room_capacity)
        # room_layout, unallocated, branch_counts_per_room, branch_range_per_room = fill_room_row_gap(pairs, room_capacity)
        # room_layout, unallocated, branch_counts_per_room, branch_range_per_room = fill_room_col_gap(pairs, room_capacity)

        ### CHANGE PATH
        build_workbook(room_layout, "C:/Users/Ankita/OneDrive/Desktop/sample.xlsx", college_name, exam_name, branch_counts_per_room, 
                      unallocated=unallocated, date="04-07-2023", shift_time="10:00-12:00", 
                      branch_range_per_room=branch_range_per_room)
        